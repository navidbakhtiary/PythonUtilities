from collections import defaultdict
from itertools import combinations
from openpyxl import load_workbook, Workbook, worksheet
from pyproj import Transformer
from pandas import DataFrame, Series
from scipy.stats import entropy, gaussian_kde, ks_2samp, mannwhitneyu
from tkinter import messagebox
import fiona
import geopandas
import glob
import numpy
import os
import pandas
import re
import time
import tkinter

date_formats = [
    "%d.%m.%Y",     # 02.08.2011
    "%d.%m.%y",     # 29.7.13
    "%Y-%m-%d %H:%M:%S",  # 2013-09-30 00:00:00
    "%d.%m.%Y.",    # 16.09.2013. (some cases have an extra dot)
    "%d-%m-%Y",     # 23-04-2012
    "%d/%m/%Y",     # 23/03/2012
    "%d.%m.%y",     # 13.08.14
    "%d %B %Y",     # 23 April 2012
]
none_values = [None, numpy.nan]

def addDataFrameAsSheetToExcel(dataframe: DataFrame, title: str, file_path: str):
    with pandas.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        dataframe.to_excel(writer, sheet_name=title, index=False)

def addLatAndLongColumnsToDataframe(dataframe: DataFrame, location_column: str = 'location', lat_column: str = 'latitude', lon_column: str = 'longitude', remove_location: bool = True):
    dataframe[[lat_column, lon_column]] = dataframe[location_column].apply(evaluateAndSplitLocation).apply(Series)
    print("Locations are splitted into Latitude and Longitude!")
    if remove_location:
        dataframe = dataframe.drop([location_column], axis='columns')
        print("column Location is removed!")
    return dataframe

def addNewDateColumnByDateRangesToDataFrame(dataframe: DataFrame, column_name: str, date_ranges: list, new_date_column_name: str, new_date_format: str):
    new_df = dataframe.copy()
    new_df[column_name] = pandas.to_datetime(new_df[column_name], format=new_date_format)
    for range in date_ranges:
        new_df.loc[(new_df[column_name] >= range[0]) & (new_df[column_name] <= range[1]), new_date_column_name] = pandas.to_datetime(range[1])
    new_df[new_date_column_name] = new_df[new_date_column_name].dt.strftime(new_date_format)
    return new_df

def addPrefixesToColumnNames(dataframe: DataFrame, column_names: list[str] = None, prefixes: list[str] | str = "df"):
    if column_names:
        if isinstance(prefixes, str):
            dataframe = dataframe.rename(columns=lambda x: prefixes + x if x in column_names else x)
        else:
            replacements_dict = dict(zip(column_names, [prefix + col for prefix, col in zip(prefixes, column_names)]))
            dataframe = dataframe.rename(columns=replacements_dict)
    elif isinstance(prefixes, str):
        dataframe = dataframe.add_prefix(prefixes)
    else:
        raise ValueError(f"Just a string prefix acceptable for all columns !")
    return dataframe

def addSuffixesToColumnNames(dataframe: DataFrame, column_names: list[str] = None, suffixes: list[str] | str = "df"):
    if column_names:
        if isinstance(suffixes, str):
            dataframe = dataframe.rename(columns=lambda x: x + suffixes if x in column_names else x)
        else:
            replacements_dict = dict(zip(column_names, [col + suffix for col, suffix in zip(column_names, suffixes)]))
            dataframe = dataframe.rename(columns=replacements_dict)
    elif isinstance(suffixes, str):
        dataframe = dataframe.add_suffix(suffixes)
    else:
        raise ValueError(f"Just a string suffix acceptable for all columns !")
    return dataframe

def calculateDegreeOfAgreement(observed: list, predicted: list):
    numerator = numpy.sum(numpy.abs(observed - predicted))
    denominator = numpy.sum(numpy.abs(observed - numpy.mean(observed)))
    d = 1 - (numerator / denominator) if denominator != 0 else 1
    return d

def calculateR2(observed: list, predicted: list):
    ss_res = numpy.sum((observed - predicted) ** 2)
    ss_tot = numpy.sum((observed - numpy.mean(observed)) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    return r2

def calculateRMSD(observed: list, predicted: list):
    rmsd = numpy.sqrt(numpy.mean((observed - predicted) ** 2))
    return rmsd

def categorizeColumnsByType(dataframe: DataFrame, keys: list = [], ignoring_columns: list = []):
    string_columns = []
    numeric_columns = []
    for header_name in [col for col in dataframe.columns if col not in keys + ignoring_columns]:
        if dataframe[header_name].apply(isNumber).all():
            numeric_columns.append(header_name)
        else:
            string_columns.append(header_name)
    return string_columns, numeric_columns

def changeDateTimeFormatInDataFrame(dataframe: DataFrame, column_names: list[str], new_formats: list[str]):
    for column_name, new_format in zip(column_names, new_formats):
        dates = dataframe[column_name].apply(convertToDatetime)
        dataframe[column_name] = dates.dt.strftime(new_format).where(dates.notna(), dataframe[column_name]).astype(str)
    return dataframe

def checkFileIsClosedBeforeSave(save_path: str):
    root = tkinter.Tk()
    root.withdraw()
    while isFileOpen(save_path):
        messagebox.showwarning("File is Open", f"The file is currently open or locked:\n\n{save_path}\n\nPlease close it and click OK to retry.")
        time.sleep(1)

def compareByBiasCorrection(observed: list, predicted: list):
    bias = numpy.mean(observed) - numpy.mean(predicted)
    corrected_predicted = predicted + bias
    difference_percent = numpy.abs(corrected_predicted - observed) / numpy.abs(observed) * 100
    return corrected_predicted, difference_percent

def compareKDE(observed: list, predicted: list):
    kde_observed = gaussian_kde(observed)
    kde_predicted = gaussian_kde(predicted)
    x = numpy.linspace(min(min(observed), min(predicted)), max(max(observed), max(predicted)), 100)
    kde_observed_vals = kde_observed(x)
    kde_predicted_vals = kde_predicted(x)
    result = abs((kde_observed_vals - kde_predicted_vals) / kde_observed_vals) * 100
    result = result[~numpy.isnan(result) & ~numpy.isinf(result)]
    difference_percent = numpy.median(result) if len(result) else None
    return difference_percent

def compareRangesDifferenceByKLDivergence(first_list: list, second_list: list):
    prob1 = numpy.histogram(first_list, bins=30, density=True)[0]
    prob2 = numpy.histogram(second_list, bins=30, density=True)[0]
    kl_div = entropy(prob1, prob2)
    difference_percent = numpy.abs(kl_div) / (numpy.mean(prob1) + numpy.mean(prob2)) * 100
    return difference_percent

def compareRangesDifferenceByKSTest(first_list: list, second_list: list):
    stat, p_value = ks_2samp(first_list, second_list)
    difference_percent = stat * 100
    return difference_percent

def compareRangesDifferenceByMannWhitney(first_list: list, second_list: list):
    stat, p_value = mannwhitneyu(first_list, second_list)
    difference_percent = numpy.abs(stat) / (numpy.mean(first_list) + numpy.mean(second_list)) * 100
    return difference_percent

def compareRangesDifferenceByQuantiles(observed: list, predicted: list, quantiles=[0.25, 0.5, 0.75]):
    observed_quantiles = numpy.quantile(observed, quantiles)
    predicted_quantiles = numpy.quantile(predicted, quantiles)
    difference_percent = abs((observed_quantiles - predicted_quantiles) / observed_quantiles) * 100
    return sum(difference_percent) / len(quantiles)

def compareValueRangesMathematically(first_list: list, second_list: list):
    observed = numpy.array(first_list)
    predicted = numpy.array(second_list)
    r2 = calculateR2(observed, predicted)
    rmsd = calculateRMSD(observed, predicted)
    degree_of_agreement = calculateDegreeOfAgreement(observed, predicted)
    rmsd_normalized = 1 / (1 + rmsd)
    total_score = (r2 + rmsd_normalized + degree_of_agreement) / 3
    print(f"R²: {r2:.4f}\n")
    print(f"RMSD: {rmsd:.4f}\n")
    print(f"Degree of Agreement (d): {degree_of_agreement:.4f}\n")
    print(f"Mean of Scores: {total_score:.4f}")
    return {'r2': r2, 'rmsd': rmsd, 'degree_of_agreement': degree_of_agreement, 'total_score': total_score}

def convertAllStringNumericsToNumeric(dataframe: DataFrame, ignoring_columns: list[str] = []):
    for column_name in [col for col in dataframe.columns if col not in ignoring_columns]:
        dataframe[column_name] = dataframe[column_name].apply(convertStringNumericToNumeric).fillna(dataframe[column_name])
    return dataframe

def convertDataFrameStringNumericToNumeric(dataframe: DataFrame, numeric_columns: list[str] = None, ignoring_columns: list[str] = None):
    if numeric_columns:
        for col in numeric_columns:
            dataframe[col] = pandas.to_numeric(dataframe[col], errors='coerce').fillna(dataframe[col])
        return dataframe
    else:
        string_headers, numeric_headers = categorizeColumnsByType(dataframe, ignoring_columns)
        for col in numeric_headers:
            try:
                dataframe[col] = pandas.to_numeric(dataframe[col])
            except Exception as exc:
                print(f"Error in converting column {col} to numeric:")
                print(exc)
        return dataframe

def convertStringNumericToNumeric(value: str):
    if isinstance(value, str) and value.isdigit():
        return int(value)
    return pandas.to_numeric(value, errors='coerce')

def convertShapeFileDataToDataFrame(file_path: str, projection_system: str):
    data = geopandas.read_file(file_path)
    if data.crs != projection_system:
        print(f"Coordinate's Projection System is changed from {data.crs} to {projection_system}")
        data = data.to_crs(projection_system)
    if all(data.geometry.geom_type == 'Point'):
        data['x'] = round(data.geometry.x, 1)
        data['y'] = round(data.geometry.y, 1)
        dataframe = DataFrame(data.drop(columns='geometry'))
        return dataframe
    else:
        return None

def convertToDatetime(value: str, source_format: str = None):
    if isinstance(value, (int, float)):
        if value > 1e12:
            return pandas.to_datetime(value, unit='ms', errors='coerce')
        elif value > 1e10:
            return pandas.to_datetime(value, unit='s', errors='coerce')
        elif 0 < value < 50000:
            return pandas.to_datetime('1899-12-30') + pandas.to_timedelta(value, unit='D')
        else:
            return pandas.NaT
    elif value not in none_values:
        if source_format:
            try:
                converted = pandas.to_datetime(value, errors='raise', format=source_format)
                print(value + " is converted to " + str(converted))
                return converted
            except ValueError:
                return numpy.nan
        else:
            for s_format in date_formats:
                try:
                    converted = pandas.to_datetime(value, errors='raise', format=s_format)
                    print(value + " is converted to " + str(converted))
                    return converted
                except ValueError:
                    continue
    return numpy.nan

def evaluateAndSplitLocation(location):
    try:
        lat, lon = map(float, location.split(', '))
        return lat, lon
    except (ValueError, AttributeError):
        return None, None

def expandColumns(dataframe: DataFrame, source_columns: list[str], destination_columns: list[str | list[str]], string_separators: list[list[str] | str], remove_source_columns: bool = False):
    df = dataframe.copy()
    re_separators = []
    for separators in string_separators:
        if isinstance(separators, list):
            pattern = "|".join(map(re.escape, separators))
        elif isinstance(separators, str):
            try:
                re.compile(separators)
                pattern = separators
            except re.error:
                pattern = re.escape(separators)
        elif isinstance(separators, re.Pattern):
            pattern = separators.pattern
        else:
            raise ValueError(f"Separator {separators} is not a list, string, or regex pattern!")
        re_separators.append(pattern)
    for s_column, d_columns, separators in zip(source_columns, destination_columns, re_separators):
        if isinstance(d_columns, str):
            d_columns = [d_columns]
        if len(d_columns) == 1:
            df[d_columns[0]] = df[s_column].str.split(separators)
            df = df.explode(d_columns[0])
        else:
            df[d_columns] = df[s_column].str.split(pattern, expand=True).replace('', pandas.NA).dropna(axis=1, how='all')
    if remove_source_columns:
        df = df.drop(columns=source_columns)
    return df

def extractCSVDataIntoDataFrame(file_path: str, file_projection_system: str, destination_projection_system: str):
    print("Extracting the data of " + file_path)
    dataframe = pandas.read_csv(file_path)
    dataframe.columns = dataframe.columns.str.strip().str.lower()
    if file_projection_system != destination_projection_system:
        print(f"Coordinate's Projection System is changed from {file_projection_system} to {destination_projection_system}")
        transformer = Transformer.from_crs(file_projection_system, destination_projection_system, always_xy=True)
        dataframe['x'], dataframe['y'] = transformer.transform(dataframe['x'].values, dataframe['y'].values)
    return dataframe

def extractDataFrame(file_path: str, sheet_names: list[str] = None, ignored_sheets: list[str] = None, headers_row_index: int = 0, first_data_row: int = 1, include_file_path: bool = False):
    try:
        dataframe = excel_data = None
        if file_path.lower().endswith('.csv'):
            dataframe = pandas.read_csv(file_path)
        elif file_path.lower().endswith(('.xls', '.xlsx')):
            if sheet_names and len(sheet_names) > 0:
                excel_data = pandas.read_excel(file_path, sheet_name=sheet_names, dtype=str)
            elif ignored_sheets and len(ignored_sheets) > 0:
                all_sheets = pandas.ExcelFile(file_path).sheet_names
                selected_sheets = [sheet for sheet in all_sheets if sheet not in ignored_sheets]
                excel_data = pandas.read_excel(file_path, sheet_name=selected_sheets, dtype=str)
            else:
                dataframe = pandas.read_excel(file_path, dtype=str)
        else:
            raise ValueError("Unsupported file type. Only .csv, .xls, or .xlsx are allowed.")
        if dataframe is not None and not dataframe.empty:
            dataframe = prepareDataFrame(dataframe, file_path, headers_row_index, first_data_row, include_file_path)
            return dataframe
        elif excel_data:
            dfs = []
            names = []
            for sheet_name, sheet_data in excel_data.items():
                sheet_data = prepareDataFrame(sheet_data, file_path, headers_row_index, first_data_row, include_file_path)
                dfs.append(sheet_data)
                names.append(sheet_name)
            if len(dfs) == 1:
                return dfs[0]
            return excel_data, dfs, names
        else:
            raise ValueError("The File is empty!")
    except Exception as e:
        return None

def fillDataFrameByAnotherDataFrame(source_dataframe: DataFrame, destination_dataframe: DataFrame, source_columns: list[str], destination_columns: list[str]):
    if len(source_columns) != len(destination_columns):
        raise ValueError("source_columns and destination_columns must have the same length")
    filled_dataframe = destination_dataframe.copy()
    for source_col, dest_col in zip(source_columns, destination_columns):
        filled_dataframe[dest_col] = source_dataframe[source_col].copy()
    return filled_dataframe

def findCSVFilesBySubstring(folder_path: str, file_name_substring: str = None):
    csv_files = []
    if file_name_substring is not None:
        csv_files = glob.glob(os.path.join(folder_path, '**', f'*{file_name_substring}*'), recursive=True)
    else:
        csv_files = glob.glob(os.path.join(folder_path, '**', '*.csv'), recursive=True)
    return csv_files

def findCSVFilesInFolder(folder_path: str, subdirectories: list = None, check_projection_system: bool = True):
    csv_files = []
    if subdirectories is not None:
        for subd in subdirectories:
            folder = os.path.join(folder_path, subd)
            csv_files.extend([os.path.abspath(file) for file in glob.glob(folder + '/**/*.csv', recursive=True)])
    else:
        csv_files = [os.path.abspath(file) for file in glob.glob(folder_path + '/**/*.csv', recursive=True)]
    if check_projection_system:
        files_info, projection_system = getDominantProjectionSystemOfCSVFiles(csv_files)
        return files_info, projection_system
    else:
        return csv_files

def findShapeFilesInFolder(folder_path: str, subdirectories: list = None):
    shp_files = []
    if subdirectories is not None:
        for subd in subdirectories:
            folder = os.path.join(folder_path, subd)
            shp_files.extend([os.path.abspath(file) for file in glob.glob(folder + '/**/*.shp', recursive=True)])
    else:
        shp_files = [os.path.abspath(file) for file in glob.glob(folder_path + '/**/*.shp', recursive=True)]
    projection_system = getDominantProjectionSystem(shp_files)
    return shp_files, projection_system

def findXLSXFilesBySubstring(folder_path: str, file_name_substring: str = None):
    if file_name_substring is not None:
        xlsx_files = glob.glob(os.path.join(folder_path, file_name_substring), recursive=True)
    else:
        xlsx_files = glob.glob(os.path.join(folder_path, '**/*.xlsx'), recursive=True)
    return xlsx_files

def freezePaneInExcelFile(workbook: Workbook, freeze_position: tuple = None):
    if freeze_position is None:
        row_to_freeze = 2
        column_to_freeze = 1
    else:
        row_to_freeze, column_to_freeze = freeze_position
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = worksheet.cell(row=row_to_freeze, column=column_to_freeze)
        print(f"Freeze panes applied to '{sheet_name}' with position {(row_to_freeze, column_to_freeze)}.")

def generateCombinations(items: list, max_count: int):
    final_list = []
    if len(items) > 1:
        for r in range(1, min(len(items), max_count)):
            combs = list(combinations(items, r))
            combs = [comb for comb in combs if len(comb) != 2 or (len(comb) == 2 and (comb[0] != comb[1]))]
            final_list.extend(combs)
        return final_list
    return []

def getDominantProjectionSystem(shape_files_path: list):
    grouped_by_projection = defaultdict(list)
    for file in shape_files_path:
        with fiona.open(file, 'r') as shapefile:
            grouped_by_projection[shapefile.crs].append(file)
    grouped_by_projection = dict(grouped_by_projection)
    for projection, file_paths in grouped_by_projection.items():
        print(f"Projection: {projection}")
        for path in file_paths:
            print(f"  File Path: {path}")
    biggest_group_projection = max(grouped_by_projection, key=lambda k: len(grouped_by_projection[k]))
    return biggest_group_projection

def getDominantProjectionSystemOfCSVFiles(csv_files_path: list):
    files_info = []
    grouped_by_projection = defaultdict(list)
    for file in csv_files_path:
        dataframe = pandas.read_csv(file)
        dataframe.columns = dataframe.columns.str.strip().str.lower()
        longitudes = dataframe['x']
        latitudes = dataframe['y']
        if longitudes.between(-180, 180).all() and latitudes.between(-90, 90).all():
            grouped_by_projection['EPSG:4326'].append(file)
            files_info.append({'file_path': file, 'projection_system': 'EPSG:4326'})
            print("Coordinates likely in geographic CRS (e.g., EPSG:4326).")
        else:
            grouped_by_projection['Other'].append(file)
            files_info.append({'file_path': file, 'projection_system': 'Other'})
            print("Coordinates likely in projected CRS.")
    grouped_by_projection = dict(grouped_by_projection)
    for projection, file_paths in grouped_by_projection.items():
        print(f"Projection: {projection}")
        for path in file_paths:
            print(f"  File Path: {path}")
    biggest_group_projection = max(grouped_by_projection, key=lambda k: len(grouped_by_projection[k]))
    return files_info, biggest_group_projection

def getNormalRangesDifference(first_list: list, second_list: list):
    first_size = abs(max(first_list) - min(first_list))
    extended_list = first_list + second_list
    extended_size = abs(max(extended_list) - min(extended_list))
    differ_percent = 100 - ((first_size / extended_size) * 100)
    return differ_percent

def getVariantRangesDifference(first_list: list, second_list: list, acceptable_percent: int = 10):
    differences = {
        'range_difference_in_percent': round(getNormalRangesDifference(first_list, second_list), 3),
        'quantile_deifference_in_percent': round(compareRangesDifferenceByQuantiles(first_list, second_list), 3),
        'KL_divergence_deifference_in_percent': round(compareRangesDifferenceByKLDivergence(first_list, second_list), 3),
        'KS_test_deifference_in_percent': round(compareRangesDifferenceByKSTest(first_list, second_list), 3),
        'mann_whitney_deifference_in_percent': round(compareRangesDifferenceByMannWhitney(first_list, second_list), 3),
    }
    return differences

def highlightDataFrameMissingValues(dataframe: DataFrame):
    styled_df = dataframe.style.apply(highlightMissingValue)
    return styled_df

def highlightMissingValue(column: Series):
    return ['background-color: yellow' if pandas.isna(value) else '' for value in column]

def insertDateByTimestampIntoDataFrame(dataframe: DataFrame, timestamp_column: str = 'timestamp', date_column_name: str = 'date'):
    timestamps = dataframe[timestamp_column].apply(convertToDatetime)
    dataframe[date_column_name] = timestamps.dt.strftime('%Y%m%d')
    return dataframe

def insertYearByTimestampIntoDataFrame(dataframe: DataFrame, timestamp_column: str = 'timestamp', year_column_name: str = 'year'):
    timestamps = dataframe[timestamp_column].apply(convertToDatetime)
    dataframe[year_column_name] = timestamps.dt.year
    return dataframe

def isNumber(value):
    if value is not None:
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False
    return True

def isFileOpen(file_path: str):
    if not os.path.exists(file_path):
        return False
    try:
        with open(file_path, 'a'):
            return False
    except IOError:
        return True

def makeAverageOnDataframe(dataframe: DataFrame, keys: list, check_numerics: bool = False, fill_missing_values: bool = True):
    print("Size of data before averaging: " + str(dataframe.shape))
    string_headers, numeric_headers = categorizeColumnsByType(dataframe, keys)
    numeric_aggregators = {header: 'mean' for header in numeric_headers}
    string_aggregators = {header: (lambda x: x.mode()[0] if not x.mode().empty else None) for header in string_headers}
    if check_numerics:
        dataframe = convertDataFrameStringNumericToNumeric(dataframe, ignoring_columns=keys)
    if fill_missing_values:
        for col in numeric_headers:
            dataframe[col] = dataframe.groupby(keys)[col].transform(lambda x: x.fillna(x.mean()))
    dataframe = dataframe.groupby(keys, as_index=False).agg(numeric_aggregators | string_aggregators)
    print("Size of data after getting Average: " + str(dataframe.shape))
    return dataframe

def makeColumnsWidthAutoFit(worksheet: worksheet):
    for column in worksheet.columns:
        max_length = len(column[0].internal_value)
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 3
        worksheet.column_dimensions[column_letter].width = adjusted_width

def makeExcelFileAutoFitWithFrozenPane(file_path: str, file_subject: str):
    workbook = load_workbook(file_path)
    makeExcelFileColumnsWidthAutoFit(workbook)
    makeExcelFileRowsHeightAutoFit(workbook)
    freezePaneInExcelFile(workbook)
    workbook.save(file_path)
    print(f"---- Result({file_subject}) is saved to: {file_path}")

def makeExcelFileColumnsWidthAutoFit(workbook: Workbook):
    for sheet_name in workbook.sheetnames:
        print("----Try to autofit the columns width of sheet: " + sheet_name)
        makeColumnsWidthAutoFit(workbook[sheet_name])

def makeExcelFileRowsHeightAutoFit(workbook: Workbook):
    for sheet_name in workbook.sheetnames:
        print("----Try to autofit the rows height of sheet: " + sheet_name)
        makeRowsHeightAutoFit(workbook[sheet_name])

def makeRowsHeightAutoFit(worksheet: worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value:
                worksheet.row_dimensions[cell.row].height = 15

def mergeDataFramesHorizontallyOnCommonColumns(dataframes: list[DataFrame], data_frames_names: list[str]):
    if len(dataframes) > 1:
        final_df = dataframes[0]
        for dataframe, df_name in zip(dataframes[1:], data_frames_names[1:]):
            common_columns = list(set(final_df.columns).intersection(set(dataframe.columns)))
            final_df = final_df.merge(dataframe, how="outer", on=common_columns, suffixes=(None, "#" + df_name))
        return final_df
    else:
        print("There is one dataframe. Nothing to merge.")

def mergeDataFramesHorizontallyOnSpecificColumns(dataframes: list[DataFrame], data_frames_names: list[str], merging_columns: list[str]):
    if len(dataframes) > 1:
        final_df = dataframes[0]
        for dataframe, df_name in zip(dataframes[1:], data_frames_names[1:]):
            final_df = final_df.merge(dataframe, how="outer", on=merging_columns, suffixes=(None, "#" + df_name))
        return final_df
    else:
        print("There is one dataframe. Nothing to merge.")

def mergeDataFramesVertically(dataframes: list[DataFrame], type_names: list[str], type_column: str, insert_index: int = 0):
    final_df = DataFrame()
    for dataframe, type_name in zip(dataframes, type_names):
        dataframe.insert(insert_index, type_column, type_name)
        final_df = pandas.concat([final_df, dataframe], ignore_index=True)
    return final_df

def mergeSheetsHorizontallyOnSpecificColumns(file_path: str, merging_columns: list[str], selected_sheets: list[str] = None):
    dfs = []
    sheet_names = []
    excel_data = extractDataFrame(file_path, selected_sheets)
    excel_data = [{'data': sheet_data, 'title': sheet_name} for sheet_name, sheet_data in excel_data.items()]
    for sheet in excel_data:
        dfs.append(sheet['data'])
        sheet_names.append(sheet['title'])
    dataframe = mergeDataFramesHorizontallyOnSpecificColumns(dfs, sheet_names, merging_columns)
    return dataframe

def mergeSheetsVertically(file_path: str, selected_sheets: list[str] = None, column_name_for_sheet_titles=None, sheet_titles: list[str] = None):
    dfs = []
    title_column_name = column_name_for_sheet_titles if column_name_for_sheet_titles else "sheet_name"
    excel_data = extractDataFrame(file_path, selected_sheets)
    if sheet_titles:
        excel_data = [{'data': sheet_data, 'title': title} for (sheet_name, sheet_data), title in zip(excel_data.items(), sheet_titles)]
    else:
        excel_data = [{'data': sheet_data, 'title': sheet_name} for sheet_name, sheet_data in excel_data.items()]
    for sheet in excel_data:
        sheet['data'][title_column_name] = sheet['title']
        dfs.append(sheet['data'])
    dataframe = pandas.concat(dfs, ignore_index=True)
    return dataframe

def normalizeDataFrame(dataframe: DataFrame, keys: list = [], ignoring_columns: list = [], variance_check: bool = True):
    temp_df = dataframe.copy()
    column_names = [col for col in temp_df.columns if col not in none_values]
    temp_df = temp_df[column_names]
    temp_df = temp_df.drop_duplicates()
    null_columns = [col for col in temp_df.columns[temp_df.isnull().all()] if col not in keys + ignoring_columns]
    temp_df = temp_df.drop(null_columns, axis=1)
    if variance_check:
        variance = temp_df.var(numeric_only=True)
        variant_columns = variance[variance > 0].index.tolist()
        cols = list(set(variant_columns + keys + ignoring_columns))
        temp_df = temp_df.loc[:, cols]
    return temp_df

def prepareDataFrame(dataframe: DataFrame, file_path: str, headers_row_index: int = 0, first_data_row: int = 1, include_file_path: bool = True):
    if headers_row_index > 0 and first_data_row > 1:
        dataframe.columns = dataframe.iloc[headers_row_index]
        dataframe = dataframe.loc[first_data_row:]
    if include_file_path:
        dataframe.insert(loc=0, column='file', value=f'=HYPERLINK("{file_path}", "{os.path.basename(file_path)}")')
    dataframe.columns = dataframe.columns.str.strip().str.lower().str.replace(' ', '_')
    dataframe = dataframe.map(lambda x: x.strip() if isinstance(x, str) else x)
    return dataframe

def reduceColumns(dataframe: DataFrame, columns_to_keep: list[str] = None, columns_to_drop: list[str] = None):
    if columns_to_drop and (not columns_to_keep):
        return dataframe.drop(columns=columns_to_drop)
    elif (not columns_to_drop) and columns_to_keep:
        columns = [col for col in dataframe.columns if col in columns_to_keep]
        return dataframe[columns]
    else:
        print("The columns are not changed! It is unclear which columns to keep!")
        return dataframe

def removeDuplicateData(dataframe: DataFrame, ignoring_columns: list = []):
    dataframe = dataframe.drop_duplicates()
    dataframe.reset_index(drop=True, inplace=True)
    duplicates = []
    columns = dataframe.columns
    for first_index in range(0, len(dataframe.columns)):
        for second_index in range(first_index + 1, len(dataframe.columns)):
            if dataframe[columns[first_index]].equals(dataframe[columns[second_index]]):
                if columns[first_index] not in ignoring_columns and columns[second_index] in ignoring_columns:
                    duplicates.append(columns[first_index])
                elif (columns[first_index] in ignoring_columns and columns[second_index] not in ignoring_columns) or (columns[first_index] not in ignoring_columns and columns[second_index] not in ignoring_columns):
                    duplicates.append(columns[second_index])
    dataframe.drop(set(duplicates), axis=1, inplace=True)
    return dataframe

def removeEmptyRows(dataframe, columns_to_check: list[str]):
    cleaned_df = dataframe.dropna(how="all", subset=columns_to_check)
    return cleaned_df

def removeSheetsFromExcelFile(file_path: str, sheet_names: list[str]):
    workbook = load_workbook(file_path)
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            workbook.remove(sheet)
    workbook.save(file_path)

def reorderColumnsOfDataFrame(dataframe: DataFrame, starter_columns: list[str], column_before_starters: str = None):
    columns = dataframe.columns.tolist()
    st_columns = [col for col in starter_columns if col in dataframe.columns]
    columns = [col for col in columns if col not in st_columns]
    new_order = None
    if column_before_starters:
        if column_before_starters in dataframe.columns:
            insert_pos = columns.index(column_before_starters) + 1
            new_order = columns[:insert_pos] + st_columns + columns[insert_pos:]
        else:
            raise ValueError(f"Column '{column_before_starters}' not found in DataFrame.")
    else:
        new_order = st_columns + columns
    return dataframe[new_order]

def replaceColumnNameOfDataFrame(dataframe: DataFrame, old_substrings: list[str], new_substrings: list[str]):
    if len(old_substrings) != len(new_substrings):
        raise ValueError("Length of old_substrings and new_substrings must be the same.")
    updated_columns = dataframe.columns.tolist()
    for old, new in zip(old_substrings, new_substrings):
        updated_columns = [col.replace(old, new) for col in updated_columns]
    dataframe.columns = updated_columns
    return dataframe

def replaceSubstringsInDataFrame(dataframe: DataFrame, column_names: list[str], old_substrings: list[list[str]], new_substrings: list[list[str]]):
    for column_name, old_subs, new_subs in zip(column_names, old_substrings, new_substrings):
        for old_sub, new_sub in zip(old_subs, new_subs):
            dataframe[column_name] = dataframe[column_name].str.replace(old_sub, new_sub, regex=False)
    return dataframe

def roundCoordinates(dataframe: DataFrame, coordinate_columns: list[str], precision_digits: list[int]):
    for index, coordinate in enumerate(coordinate_columns):
        dataframe[coordinate] = pandas.to_numeric(dataframe[coordinate], errors='coerce').round(precision_digits[index])
    return dataframe


def saveDataFramesInExcel(dataframes: list[DataFrame], sheet_names: list[str], save_path: str, file_subject: str = "", freeze_position: tuple = None):
    checkFileIsClosedBeforeSave(save_path)
    with pandas.ExcelWriter(save_path, engine='openpyxl') as writer:
        for dataframe, sheet_name in zip(dataframes, sheet_names):
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
    makeExcelFileAutoFitWithFrozenPane(save_path, file_subject)


def saveToExcel(dataframe: DataFrame, save_path: str, file_subject: str = ""):
    checkFileIsClosedBeforeSave(save_path)
    dataframe.to_excel(save_path, index=False)
    makeExcelFileAutoFitWithFrozenPane(save_path, file_subject)


def splitDataFrameHorizontally(dataframe: DataFrame, common_columns: list[str], columns_to_split: list[str]):
    dfs = []
    for column in columns_to_split:
        df = dataframe[common_columns + [column]]
        df = df[df[column].notna() & (df[column] != '')]
        if not df.empty:
            dfs.append(df)
    return dfs


def splitDataFrameVertically(dataframe: DataFrame, grouping_columns: list[str]):
    grouped_data = dataframe.groupby(grouping_columns, as_index=False)
    grouped_result = []
    for group_values, group_df in grouped_data:
        grouped_result.append({'values': group_values, 'data': group_df})
    return grouped_result


def splitDataFrameVerticallyIntoExcelFiles(dataframe: DataFrame, grouping_columns: list[str], save_folder: str, file_name_prefix: str = None, data_value_as_file_name: bool = True):
    os.makedirs(save_folder, exist_ok=True)
    files_pathes = []
    grouped_data = splitDataFrameVertically(dataframe, grouping_columns)
    counter = 1
    for group in grouped_data:
        group_values = group["values"]
        group_df = group["data"]
        if not isinstance(group_values, tuple):
            group_values = (group_values,)
        if data_value_as_file_name:
            file_name = file_name_prefix + "_".join(str(value) for value in group_values) + ".xlsx"
        else:
            file_name = file_name_prefix + "_" + str(counter) + ".xlsx"
            counter += 1
        file_path = os.path.join(save_folder, file_name)
        files_pathes.append(file_path)
        saveToExcel(group_df, file_path, file_subject=str(group_values))

    return files_pathes


def uniqueValuesCount(values):
    return values.dropna().nunique()
